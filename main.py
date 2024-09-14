from Jarvis import JarvisAssistant
import re
import os
import random
import pprint
import datetime
import requests
import sys
import pyjokes
import time
import pyautogui
import pywhatkit
import webbrowser
import winshell
import wolframalpha
import subprocess
from openpyxl import load_workbook
from PIL import Image
from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtCore import QTimer, QTime, QDate, Qt
from PyQt5.QtGui import QMovie
from PyQt5.QtWidgets import QApplication, QMainWindow
from PyQt5.uic import loadUiType
from Jarvis.features.gui import Ui_MainWindow
from Jarvis.config import config
from PyQt5.QtCore import QThread


obj = JarvisAssistant()

# ================================ MEMORY ===========================================================================================================

GREETINGS = ["hello jarvis", "jarvis", "wake up jarvis", "you there jarvis", "time to work jarvis", "hey jarvis",
             "ok jarvis", "are you there jarvis"]
GREETINGS_RES = ["always there for you sir", "i am ready sir",
                 "your wish is my command sir", "how can i help you sir?", "i am online and ready sir"]

EMAIL_DIC = {
    'myself': 'atharvaaingle@gmail.com',
    'my official email': 'esamba333@gmail.com',
    'my second email': 'esamba333@gmail.com',
    'my official mail': 'esamba333@gmail.com',
    'my second mail': 'esamba333@gmail.com'
}

CALENDAR_STRS = ["what do i have", "do i have plans", "am i busy"]
# =======================================================================================================================================================


def speak(text):
    obj.tts(text)


app_id = config.wolframalpha_id


def computational_intelligence(question):
    try:
        client = wolframalpha.Client(app_id)
        answer = client.query(question)
        answer = next(answer.results).text
        print(answer)
        return answer
    except Exception as e:
        speak("Sorry sir I couldn't fetch your question's answer. Please try again ")
        print(e)
        return None


def startup():
    speak("Initializing Jarvis")
    speak("Starting all systems applications")
    speak("Installing and checking all drivers")
    speak("Calibrating and examining all the core processors")
    speak("Checking the internet connection")
    speak("Wait a moment sir")
    speak("All drivers are up and running")
    speak("All systems have been activated sir")
    speak("Now I am online")


def wish():
    hour = int(datetime.datetime.now().hour)
    if hour >= 0 and hour <= 12:
        speak("Good Morning")
    elif hour > 12 and hour < 18:
        speak("Good afternoon")
    else:
        speak("Good evening")
    c_time = obj.tell_time()
    speak(f"Currently it is {c_time}")
    speak("I am Jarvis. Online and ready sir. Please tell me how may I help you")


def load_excel_data():
    try:
        file_path = "C:/Users/otien/Documents/inputs.xlsx"
        wb = load_workbook(file_path)

        # Load commands from the 'User' sheet
        ws_user = wb['User']
        max_col_user = ws_user.max_column  # Get the maximum number of columns
        max_row_user = ws_user.max_row  # Get the maximum number of rows

        # Initialize lists to store data
        hello_list = []
        how_are_you = []
        # You can add more lists here for additional columns if needed

        # Read data from each column
        for row in range(1, max_row_user + 1):  # Assuming you need all rows
            for col in range(1, max_col_user + 1):
                cell_value = ws_user.cell(row=row, column=col).value
                if cell_value:
                    if col == 1:  # Column 1 for hello_list
                        hello_list.append(cell_value.strip().lower())
                    elif col == 2:  # Column 2 for how_are_you
                        how_are_you.append(cell_value.strip().lower())
                    # Add additional conditions for other columns if needed

        # Load replies from the 'Replies' sheet
        ws_replies = wb['Replies']
        max_col_replies = ws_replies.max_column  # Get the maximum number of columns
        max_row_replies = ws_replies.max_row  # Get the maximum number of rows

        reply_hello_list = []
        reply_how_are_you = []
        # You can add more lists here for additional columns if needed

        # Read data from each column
        for row in range(1, max_row_replies + 1):  # Assuming you need all rows
            for col in range(1, max_col_replies + 1):
                cell_value = ws_replies.cell(row=row, column=col).value
                if cell_value:
                    if col == 1:  # Column 1 for reply_hello_list
                        reply_hello_list.append(cell_value.strip())
                    elif col == 2:  # Column 2 for reply_how_are_you
                        reply_how_are_you.append(cell_value.strip())
                    # Add additional conditions for other columns if needed

        return hello_list, reply_hello_list, how_are_you, reply_how_are_you

    except Exception as e:
        speak(f"Error loading Excel file: {e}")
        print(f"Error loading Excel file: {e}")
        return [], [], [], []
    
def close_notepad():
    try:
        subprocess.run(["taskkill", "/IM", "notepad.exe", "/F"], check=True)
        speak("Notepad closed successfully.")
    except subprocess.CalledProcessError as e:
        speak(f"Failed to close Notepad. Error: {e}")

def close_vscode():
    try:
        subprocess.run(["taskkill", "/IM", "Code.exe", "/F"], check=True)
        speak("Visual Studio Code closed successfully.")
    except subprocess.CalledProcessError as e:
        speak(f"Failed to close Visual Studio Code. Error: {e}")        

def show_info():
    speak("Fetching system information...")
    info = {
        "Operating System": os.uname().sysname,
        "OS Version": os.uname().version,
        "Machine": os.uname().machine,
        "Processor": os.uname().processor,
        "System Name": os.uname().nodename
    }
    
    for key, value in info.items():
        speak(f"{key}: {value}")

class MainThread(QThread):
    def __init__(self):
        super(MainThread, self).__init__()

    def run(self):
        self.TaskExecution()

    def TaskExecution(self):
        startup()
        wish()
        hello_list, reply_hello_list, how_are_you, reply_how_are_you = load_excel_data()

        while True:
            command = obj.mic_input().strip().lower()  # Ensure command is in lowercase
            print(f"Command Received: {command}")  # Debug print

            if re.search('date', command):
                date = obj.tell_me_date()
                print(date)
                speak(date)

            elif any(keyword in command for keyword in ["what's the time", "tell me the time", "what's the time", "jarvis what's the time"]):
                time_c = obj.tell_time()
                print(time_c)
                speak(f"Sir, the time is {time_c}")

            elif command in hello_list:
                index = hello_list.index(command)
                response = reply_hello_list[index] if index < len(reply_hello_list) else "Sorry, I don't have a response for that."
                print(f"Matched Greeting Command: {response}")  # Debug print
                speak(response)

            elif command in how_are_you:
                index = how_are_you.index(command)
                response = reply_how_are_you[index] if index < len(reply_how_are_you) else "Sorry, I don't have a response for that."
                print(f"Matched How Are You Command: {response}")  # Debug print
                speak(response)

            elif re.search('launch', command):
                dict_app = {
                    'chrome': 'C:/Program Files/Google/Chrome/Application/chrome.exe',
                    'visual studio': 'C:\\Users\\otien\\AppData\\Local\\Programs\\Microsoft VS Code\\Code.exe',
                    'cmd': 'C:\\Windows\\system32\\cmd.exe',
                    'notepad' : 'C:\\Windows\\system32\\notepad.exe'
                }

                app = command.split(' ', 1)[1]
                path = dict_app.get(app)

                if path is None:
                    speak('Application path not found')
                    print('Application path not found')
                else:
                    speak(f'Launching: {app} for you sir!')
                    obj.launch_any_app(path_of_app=path)

            elif command in GREETINGS:
                speak(random.choice(GREETINGS_RES))

            elif re.search('open', command):
                domain = command.split(' ')[-1]
                open_result = obj.website_opener(domain)
                speak(f'Alright sir! Opening {domain}')
                print(open_result)

            elif re.search('weather', command):
                city = command.split(' ')[-1]
                weather_res = obj.weather(city=city)
                print(weather_res)
                speak(weather_res)

            elif 'open amazon' in command: 
                speak("Opening Amazon. Happy shopping!")
                webbrowser.open("https://www.amazon.com") 

            elif 'fetch system information' in command:
                show_info()        

            elif 'shutdown' in command or 'shutdown my pc' in command:
                speak("Your system is about to shut down. Please save your work.")
                time.sleep(5)
                subprocess.call(['shutdown', '/s'])

            elif 'restart' in command or 'restart my pc javis' in command or 'restart my pc' in command:
                speak("Restarting the PC.")
                subprocess.call(['shutdown', '/r'])
                
            elif 'hibernate' in command or 'hibernate my pc ' in command:
                speak("Hibernating the system.")
                subprocess.call(['shutdown', '/h'])

            elif 'log off' in command  or 'logout the pc ' in command:
                speak("Signing out. Please save your work.")
                time.sleep(5)
                subprocess.call(['shutdown', '/l'])    

            elif re.search('tell me about', command):
                topic = command.split(' ')[-1]
                if topic:
                    wiki_res = obj.tell_me(topic)
                    print(wiki_res)
                    speak(wiki_res)
                else:
                    speak("Sorry sir. I couldn't load your query from my database. Please try again")

            elif any(keyword in command for keyword in ["buzzing", "news", "headlines", "update me", "what's cooking"]):
                news_res = obj.news()
                speak('Source: The Nairobian')
                speak('Today\'s Headlines are..')
                for index, articles in enumerate(news_res):
                    pprint.pprint(articles['title'])
                    speak(articles['title'])
                    if index == len(news_res) - 2:
                        break
                speak('These were the top headlines. Have a nice day sir!')

            elif 'search google for' in command:
                obj.search_anything_google(command)

            elif any(keyword in command for keyword in ["play music", "hit some music"]):
                music_dir = "C:\\Users\\otien\\Downloads\\Dj numz"
                songs = os.listdir(music_dir)
                for song in songs:
                    os.startfile(os.path.join(music_dir, song))

            elif 'open youtube' in command:
                video = command.split(' ')[1]
                speak(f"Okay sir, playing {video} on YouTube")
                pywhatkit.playonyt(video)

            elif any(keyword in command for keyword in ["who created you?", "who created you jarvis?", "who is samba?"]):
                speak("Samba is my developer. He developed me to assist with tasks such as playing music, opening browsers, and more.")

            elif any(keyword in command for keyword in ["email", "send email"]):
                sender_email = config.email
                sender_password = config.email_password

                try:
                    speak("Whom do you want to email, sir?")
                    recipient = obj.mic_input()
                    receiver_email = EMAIL_DIC.get(recipient)
                    if receiver_email:
                        speak("What is the subject, sir?")
                        subject = obj.mic_input()
                        speak("What should I say?")
                        message = obj.mic_input()
                        msg = f'Subject: {subject}\n\n{message}'
                        obj.send_mail(sender_email, sender_password, receiver_email, msg)
                        speak("Email has been successfully sent")
                        time.sleep(2)
                    else:
                        speak("I couldn't find the requested person's email in my database. Please try again with a different name.")
                except Exception as e:
                    speak("Sorry sir. Couldn't send your mail. Please try again")

            elif 'how are you' in command:
                speak("I'm fine, thank you sir. How are you sir?")

            elif "thank you" in command:
                speak("The pleasure is all mine. Anything that i can help you with?")

            elif any(keyword in command for keyword in ['i am fine', 'i am good', 'i am fine thank you']):
                speak("Good to hear that!")

            elif "calculate" in command or "what is" in command or "who is" in command:
                question = command
                answer = computational_intelligence(question)
                speak(answer)

            elif any(keyword in command for keyword in ["what do i have", "do i have plans", "am i busy"]):
                obj.google_calendar_events(command)

            elif 'close visual studio' in command:
                  close_vscode()

            elif 'close notepad' in command:
                   close_notepad()      
            elif any(keyword in command for keyword in ["make a note", "write this down", "remember this",'take note','take note of these']):
                speak("What would you like me to write down?")
                note_text = obj.mic_input()
                obj.take_note(note_text)
                speak("I've made a note of that")

            elif any(keyword in command for keyword in ["close the note", "close notepad"]):
                speak("Okay sir, closing Notepad")
                os.system("taskkill /f /im notepad++.exe")

            elif any(keyword in command for keyword in ["notepad", "open notepad"]):
                obj.note()

            elif any(keyword in command for keyword in ["give me a joke", "tell me a joke"]):
                joke = pyjokes.get_joke()
                print(joke)
                speak(joke)

            elif "give me the system information" in command:
                sys_info = obj.system_info()
                print(sys_info)
                speak(sys_info)

            elif any(keyword in command for keyword in ["where is", "locate"]):
                place = command.split('where is ', 1)[1]
                current_loc, target_loc, distance = obj.location(place)
                city = target_loc.get('city', '')
                state = target_loc.get('state', '')
                country = target_loc.get('country', '')
                time.sleep(1)
                try:
                    if city:
                        res = f"{place} is in {state} state and country {country}. It is {distance} km away from your current location."
                        print(res)
                        speak(res)
                    else:
                        res = f"{state} is a state in {country}. It is {distance} km away from your current location."
                        print(res)
                        speak(res)
                except Exception as e:
                    res = "Sorry sir, I couldn't get the coordinates of the location you requested. Please try again."
                    speak(res)

            elif "ip address" in command:
                ip = requests.get('https://api.ipify.org').text
                print(ip)
                speak(f"Your IP address is {ip}")

            elif any(keyword in command for keyword in ["switch the window", "switch window"]):
                speak("Okay sir, switching the window")
                pyautogui.keyDown("alt")
                pyautogui.press("tab")
                time.sleep(1)
                pyautogui.keyUp("alt")

            elif any(keyword in command for keyword in ["where am i", "current location", "where i am"]):
                try:
                    city, state, country = obj.my_location()
                    print(city, state, country)
                    speak(f"You are currently in {city}, which is in {state} state and country {country}.")
                except Exception as e:
                    speak("Sorry sir, I couldn't fetch your current location. Please try again.")

            elif any(keyword in command for keyword in ["take screenshot", "take a screenshot", "capture the screen"]):
                speak("By what name do you want to save the screenshot?")
                name = obj.mic_input()
                speak("Alright sir, taking the screenshot")
                img = pyautogui.screenshot()
                name = f"{name}.png"
                img.save(name)
                speak("The screenshot has been successfully captured")

            elif "show me the screenshot" in command:
                try:
                    img = Image.open(f'C://Users//otien//Desktop//JARVIS-MASTER//{name}')
                    img.show()
                    speak("Here it is sir")
                    time.sleep(2)
                except IOError:
                    speak("Sorry sir, I am unable to display the screenshot")

            elif any(keyword in command for keyword in ["hide all files", "hide this folder"]):
                os.system("attrib +h /s /d")
                speak("Sir, all the files in this folder are now hidden")

            elif any(keyword in command for keyword in ["visible", "make files visible"]):
                os.system("attrib -h /s /d")
                speak("Sir, all the files in this folder are now visible to everyone. I hope you are taking this decision in your own peace.")

            elif 'empty the recycle bin' in command:
                speak("Are you sure about this, sir?")
                my_answer = obj.mic_input()  # Fixed: obj.mic_input() should be used to get the answer
                if 'yes' in my_answer:
                    winshell.recycle_bin().empty(confirm="False", show_progress="False", sound="True")
                    speak("Recycle bin cleaned.")
                else:
                    speak("The data in the recycle bin is still there.")

            elif any(keyword in command for keyword in ["search", "open"]) and any(keyword in command for keyword in [" in youtube", "google"]):
                speak("What do you want to search for, sir?")
                answer = obj.mic_input()  # Fixed: obj.mic_input() should be used to get the search query
                search_url = "https://www.youtube.com/search?q=" + answer
                webbrowser.open(search_url)

            elif "play" in command and "song" in command and "in youtube" in command:
                speak("What song do you want to play?")
                song_name = obj.mic_input()  # Fixed: obj.mic_input() should be used to get the song name

                # Replace spaces with '+' for URL compatibility
                song_name = song_name.replace(" ", "+")
                search_url = "https://www.youtube.com/results?search_query=" + song_name
                webbrowser.open(search_url)

            elif any(keyword in command for keyword in ["goodbye jarvis", "go offline", "bye",'exit']):
                speak("Alright sir, going offline. It was nice working with you")
                sys.exit()


startExecution = MainThread()


class Main(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.ui.pushButton.clicked.connect(self.startTask)
        self.ui.pushButton_2.clicked.connect(self.close)

    def __del__(self):
        sys.stdout = sys.__stdout__

    def startTask(self):
        self.ui.movie = QtGui.QMovie("Jarvis/utils/images/live_wallpaper.gif")
        self.ui.label.setMovie(self.ui.movie)
        self.ui.movie.start()
        self.ui.movie = QtGui.QMovie("Jarvis/utils/images/initiating.gif")
        self.ui.label_2.setMovie(self.ui.movie)
        self.ui.movie.start()
        timer = QTimer(self)
        timer.timeout.connect(self.showTime)
        timer.start(1000)
        startExecution.start()

    def showTime(self):
        current_time = QTime.currentTime()
        current_date = QDate.currentDate()
        label_time = current_time.toString('hh:mm:ss')
        label_date = current_date.toString(Qt.ISODate)
        self.ui.textBrowser.setText(label_date)
        self.ui.textBrowser_2.setText(label_time)


app = QApplication(sys.argv)
jarvis = Main()
jarvis.show()
sys.exit(app.exec_())
